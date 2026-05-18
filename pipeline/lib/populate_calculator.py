#!/usr/bin/env python3
"""
populate_calculator.py — Universal LEED calculator population engine.

Usage (CLI):
    python3 populate_calculator.py <template_path> <credit_code> <data_json_path> <output_path>

Prints JSON result to stdout. Prints progress to stderr.
Exit 0 on success, 1 on error.

Function:
    populate_calculator(template_path, credit_code, data_json, output_path) -> dict

Does NOT call any AI. Receives structured data_json (produced upstream by Claude)
and writes it into the USGBC Excel template using openpyxl. Works for every
calculator — discovers structure by reading the template. Never hardcodes cell
references or calculator-specific logic.
"""

import sys
import json
import copy
import re
import shutil
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
except ImportError:
    print(json.dumps({"error": "openpyxl not installed — run: pip3 install openpyxl"}))
    sys.exit(1)

SCHEMA_PATH = (
    Path(__file__).resolve().parent.parent
    / "reference" / "leed" / "leed_v41_calculator_schemas.json"
)


# ── Merged cell helper ─────────────────────────────────────────────────────────

def writable_cell(ws, row: int, col: int):
    """
    Return the writable cell for (row, col). openpyxl raises AttributeError when
    you try to set .value on a MergedCell that is not the anchor (top-left) of
    its merge range. This resolves to the anchor instead.
    """
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return cell
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(row=rng.min_row, column=rng.min_col)
    return cell  # fallback


# ── Schema loading ─────────────────────────────────────────────────────────────

def _cc(s: str) -> str:
    """Collapse a string to lowercase alphanumeric for loose comparison."""
    return re.sub(r"[^a-z0-9]", "", s.lower())


def load_schema(credit_code: str) -> dict:
    """Return the schema dict for the given credit code (flexible matching)."""
    with open(SCHEMA_PATH) as f:
        all_schemas = json.load(f)

    calculators = all_schemas.get("calculators", {})
    cc = _cc(credit_code)

    for key, schema in calculators.items():
        if _cc(key) == cc or _cc(schema.get("id", "")) == cc:
            return schema

    for schema in calculators.values():
        for credit in schema.get("credits", []):
            if _cc(credit) == cc:
                return schema

    raise ValueError(f"No calculator schema found for credit_code: {credit_code!r}")


# ── Label matching ─────────────────────────────────────────────────────────────

def _normalize(text: str) -> str:
    """Strip noise from a cell label for matching."""
    s = str(text).strip().lower()
    s = re.sub(r"\s*[\(\[][^)\]]*[\)\]]", "", s)   # remove (...) and [...]
    s = re.sub(r"[:\*†‡—–\-]+$", "", s)             # trailing punctuation
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _match_score(a: str, b: str) -> float:
    """Similarity score 0–1 between two labels."""
    na, nb = _normalize(a), _normalize(b)
    if not na or not nb:
        return 0.0
    if na == nb:
        return 1.0
    # Only do substring matching when both sides are meaningful phrases (≥ 3 chars).
    # A single letter like 'A' must NOT score 0.9 against 'Space / Zone Name'.
    if len(na) >= 3 and len(nb) >= 3 and (na in nb or nb in na):
        return 0.9
    wa = set(na.split())
    wb_ = set(nb.split())
    # Skip word-overlap when one side is a single short token — avoids false positives
    # from abbreviations ('Az', 'Pz', 'A', 'B') matching real field names.
    if wa and wb_ and min(len(na), len(nb)) >= 3:
        return len(wa & wb_) / max(len(wa), len(wb_))
    return 0.0


def find_input_cell(ws, label: str, max_rows: int = 150):
    """
    Scan the sheet for a cell whose text matches `label`.
    Returns (row, col) of the input cell, or None.

    USGBC calculators often have: [label] [abbreviation] [units formula] [INPUT]
    so we scan up to 6 cells to the right of the label looking for the first
    truly empty (None) cell, skipping over abbreviations and unit-formula cells.
    """
    best_score = 0.62
    best_loc = None
    max_col = min((ws.max_column or 1) + 1, 40)

    for r in range(1, min((ws.max_row or 1) + 1, max_rows)):
        for c in range(1, max_col):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            if not cell.value or not isinstance(cell.value, str):
                continue
            score = _match_score(cell.value, label)
            if score <= best_score:
                continue

            # Scan to the right for the first empty (None) cell — that is the input
            found_loc = None
            for offset in range(1, 7):
                if c + offset > max_col:
                    break
                candidate = writable_cell(ws, r, c + offset)
                cand_val = candidate.value
                if cand_val is None:
                    # First empty cell to the right — this is the input cell
                    found_loc = (candidate.row, candidate.column)
                    break
                if isinstance(cand_val, str) and cand_val.startswith("="):
                    # Formula cell — skip it (it's a calculated column)
                    continue
                # Non-empty, non-formula → abbreviation or unit label — skip it

            if found_loc:
                best_score = score
                best_loc = found_loc

    return best_loc


# ── Header-row detection ───────────────────────────────────────────────────────

def find_header_row(ws, field_names: list, max_rows: int = 80):
    """
    Find the row whose cells best match the given column-header names.
    Returns (row_index, {field_name: col_index}) or None.
    Requires at least 2 matching columns (or ≥ ⌊len/3⌋).
    Each field is assigned to the HIGHEST-scoring column (not the first ≥ 0.7).
    Ties broken by cumulative match score, then by earliest row.
    """
    best_row, best_count, best_qual, best_map = None, 0, 0.0, {}
    min_match = max(2, len(field_names) // 3)
    max_col = min((ws.max_column or 1) + 1, 40)

    for r in range(1, min((ws.max_row or 1) + 1, max_rows)):
        # For each field, find the BEST (highest-score) matching column in this row.
        per_field_best: dict[str, tuple[float, int]] = {}  # field → (score, col)
        for c in range(1, max_col):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell) or not cell.value:
                continue
            for field in field_names:
                s = _match_score(str(cell.value), field)
                if s >= 0.7:
                    prev_s, _ = per_field_best.get(field, (0.0, -1))
                    if s > prev_s:
                        per_field_best[field] = (s, c)

        mapping = {f: c for f, (s, c) in per_field_best.items()}
        scores  = {f: s for f, (s, c) in per_field_best.items()}
        n = len(mapping)
        qual = sum(scores.values())
        if n >= min_match and (n > best_count or (n == best_count and qual > best_qual)):
            best_count, best_qual, best_row, best_map = n, qual, r, mapping

    return (best_row, best_map) if best_row else None


# ── Template row analysis ──────────────────────────────────────────────────────

def find_data_row_range(ws, header_row: int, col_map: dict = None):
    """
    After a header row find the range of pre-built template data rows.

    USGBC calculators often have sub-header rows immediately after the column
    header row (e.g., abbreviation rows like Az/Pz, unit rows like (sq ft)).
    Key insight: sub-header rows always have SHORT TEXT in the mapped input
    columns (e.g. 'Az', 'Pz', '(people)'). Actual data rows have ALL input
    columns empty (None) or formula — they are waiting for user input.

    Returns (first_data_row, last_data_row).
    """
    input_cols = list(col_map.values()) if col_map else []
    first = None
    last = None

    for r in range(header_row + 1, min((ws.max_row or header_row + 1) + 1, header_row + 80)):
        row_vals = [
            ws.cell(row=r, column=c).value
            for c in range(1, 25)
            if not isinstance(ws.cell(row=r, column=c), MergedCell)
        ]
        non_empty = [v for v in row_vals if v is not None]

        if not non_empty:
            if first is not None:
                break  # blank row after data section — done
            continue  # blank row before data section — keep scanning

        text_vals = [str(v).strip().lower() for v in non_empty if isinstance(v, str)]
        if any(kw in t for kw in ("total", "sum", "subtotal", "overall", "result") for t in text_vals):
            break  # summary/total row — stop before it

        if input_cols:
            input_vals = [ws.cell(row=r, column=c).value for c in input_cols]

            # ALL input columns are None or formula → actual data row (ready for entry)
            all_input_empty = all(
                v is None or (isinstance(v, str) and v.startswith("="))
                for v in input_vals
            )

            if not all_input_empty:
                # At least one input column has a value — check if it's a real data value
                # (number or long string) vs a label/abbreviation/unit (short string).
                has_real_data = any(
                    v is not None
                    and not (isinstance(v, str) and v.startswith("="))
                    and (
                        isinstance(v, (int, float))
                        or (isinstance(v, str) and len(v.strip()) > 20)
                    )
                    for v in input_vals
                )
                if not has_real_data:
                    # Short text in input cols (e.g. 'Az', 'Pz', '(people)') → sub-header
                    if first is None:
                        continue  # haven't started yet — skip this sub-header
                    else:
                        break  # past data section — stop

        if first is None:
            first = r
        last = r

    if first is None:
        first = header_row + 1
        last = first
    return (first, max(last, first))


# ── Formula adjustment ─────────────────────────────────────────────────────────

def shift_formula(formula: str, src_row: int, dst_row: int) -> str:
    """
    Shift relative row references in a formula when copying from src_row to dst_row.
    Absolute references ($ROW) are not changed.
    """
    if not formula or not isinstance(formula, str) or not formula.startswith("="):
        return formula
    offset = dst_row - src_row
    if offset == 0:
        return formula

    def replace_ref(m):
        col = m.group(1)
        row_num = int(m.group(2))
        # If preceded by $, it is an absolute row ref — do not shift
        start = m.start()
        if start > 0 and formula[start - 1] == "$":
            return m.group(0)
        return f"{col}{row_num + offset}"

    return re.sub(r"([A-Za-z]+)(\d+)", replace_ref, formula)


def expand_ranges_in_sheet(ws, old_last: int, new_last: int):
    """
    After inserting rows, extend SUM/AVERAGE ranges that ended at old_last
    to end at new_last instead.
    """
    if old_last == new_last:
        return

    def expand(m):
        sc, sr, ec, er = m.group(1), m.group(2), m.group(3), int(m.group(4))
        if er == old_last:
            return f"{sc}{sr}:{ec}{new_last}"
        return m.group(0)

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                new_val = re.sub(r"([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)", expand, cell.value)
                if new_val != cell.value:
                    cell.value = new_val


# ── Row copying ────────────────────────────────────────────────────────────────

def copy_row(ws, src_row: int, dst_row: int):
    """
    Copy formatting and formulas (with row-shifted references) from src_row to dst_row.
    Input cells (non-formula) are left blank so data can be written into them.
    Skips MergedCell sub-cells — only writes to anchor cells.
    """
    max_col = (ws.max_column or 1) + 1
    for c in range(1, max_col):
        src = ws.cell(row=src_row, column=c)
        if isinstance(src, MergedCell):
            continue
        dst = writable_cell(ws, dst_row, c)
        if isinstance(dst, MergedCell):
            continue
        if src.has_style:
            dst.font = copy.copy(src.font)
            dst.border = copy.copy(src.border)
            dst.fill = copy.copy(src.fill)
            dst.alignment = copy.copy(src.alignment)
            dst.number_format = src.number_format
        # ArrayFormula objects cannot be safely copied by openpyxl — skip them.
        # The original template rows retain the array formula; inserted rows
        # will inherit the calculation via Excel's own recalculation on open.
        try:
            from openpyxl.worksheet.formula import ArrayFormula
            if isinstance(src.value, ArrayFormula):
                dst.value = None
                continue
        except ImportError:
            pass
        if isinstance(src.value, str) and src.value.startswith("="):
            dst.value = shift_formula(src.value, src_row, dst_row)
        else:
            dst.value = None


# ── Occupancy category lookup ──────────────────────────────────────────────────

def load_occupancy_categories(wb) -> list:
    """
    Extract the list of valid occupancy category strings from the Instructions sheet.
    These are the exact strings the VLOOKUP formula requires — they typically have
    trailing spaces (e.g. 'Office space ', 'Lobbies ').

    Locates the 'Occupancy Category' header in the sheet, then collects all
    non-formula string values in the same column below it.
    """
    if "Instructions" not in wb.sheetnames:
        return []
    ws = wb["Instructions"]

    # Step 1: find the row of the "Occupancy Category" column header
    header_row = None
    header_col = None
    for r in range(1, min((ws.max_row or 1) + 1, 120)):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell) or not cell.value:
                continue
            if isinstance(cell.value, str) and "occupancy" in cell.value.lower() and "category" in cell.value.lower():
                header_row = r
                header_col = c
                break
        if header_row:
            break

    if not header_row:
        return []

    # Step 2: collect all non-formula strings in that column below the header
    cats = []
    for r in range(header_row + 1, min((ws.max_row or header_row + 1) + 1, header_row + 120)):
        cell = ws.cell(row=r, column=header_col)
        if isinstance(cell, MergedCell) or not cell.value:
            continue
        v = cell.value
        if not isinstance(v, str) or v.startswith("="):
            continue
        stripped = v.strip()
        if len(stripped) < 3:
            continue
        cats.append(v)

    return cats


def snap_occupancy_category(value: str, valid_cats: list) -> str:
    """
    Map Claude's approximate occupancy category string to the nearest valid entry.
    Uses normalized scoring. Falls back to the original value if no good match.
    """
    if not valid_cats:
        return value
    val_norm = _normalize(value)
    best_score, best_cat = 0.0, value
    for cat in valid_cats:
        score = _match_score(value, cat)
        if score > best_score:
            best_score, best_cat = score, cat
    # Only snap if reasonably confident
    return best_cat if best_score >= 0.6 else value


# ── Static field writing ───────────────────────────────────────────────────────

def write_static_fields(ws, static_fields: dict, result: dict):
    """Write label→value pairs by scanning the sheet for matching label cells."""
    for label, value in static_fields.items():
        if value is None:
            continue
        loc = find_input_cell(ws, label)
        if loc:
            row, col = loc
            cell = writable_cell(ws, row, col)
            if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.value = value
                result["cells_written"] += 1
        else:
            result["warnings"].append(f"No input cell found for label: {label!r}")


# ── Dynamic row writing ────────────────────────────────────────────────────────

def write_rows(ws, rows: list, result: dict, occupancy_cats: list = None):
    """
    Write a list of row-dicts into a tabular section of the sheet.
    Inserts and formats new rows when the template doesn't have enough.
    """
    if not rows:
        return

    field_names = list(rows[0].keys())
    found = find_header_row(ws, field_names)
    if not found:
        result["warnings"].append(f"Header row not found for fields: {field_names[:4]}")
        return

    header_row, col_map = found
    first_data, last_template = find_data_row_range(ws, header_row, col_map)
    n_template = last_template - first_data + 1
    n_needed = len(rows)

    # Insert extra rows when template doesn't have enough
    if n_needed > n_template:
        extra = n_needed - n_template
        insert_at = last_template + 1
        ws.insert_rows(insert_at, amount=extra)
        for i in range(extra):
            copy_row(ws, last_template, insert_at + i)
        expand_ranges_in_sheet(ws, last_template, last_template + extra)
        last_template += extra

    # Write values into each row
    for i, row_data in enumerate(rows):
        r = first_data + i
        for field, value in row_data.items():
            if value is None:
                continue
            col = col_map.get(field)
            if col is None:
                # Lazy lookup: try to find the column now
                for c in range(1, min((ws.max_column or 1) + 1, 40)):
                    hdr = ws.cell(row=header_row, column=c).value
                    if hdr and _match_score(str(hdr), field) >= 0.7:
                        col = c
                        col_map[field] = c
                        break
            if col:
                cell = writable_cell(ws, r, col)
                if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                    write_val = value
                    # Snap occupancy category to the exact string required by VLOOKUP
                    if (
                        occupancy_cats
                        and isinstance(value, str)
                        and "occupanc" in field.lower()
                    ):
                        write_val = snap_occupancy_category(value, occupancy_cats)
                    cell.value = write_val
                    result["cells_written"] += 1
            else:
                result["warnings"].append(f"No column found for field: {field!r}")


# ── Systems mode ───────────────────────────────────────────────────────────────

def write_systems(ws, systems: list, result: dict, occupancy_cats: list = None):
    """
    For calculators with multiple system blocks (each system has header fields
    and a zone/fixture table). The first system uses the template's existing
    header and data rows. All items from all systems are written as a flat
    sequence into the tab's repeating row section.
    """
    if not systems:
        return

    # Write first system's static fields into the template's header section
    if systems[0].get("static_fields"):
        write_static_fields(ws, systems[0]["static_fields"], result)

    # Collect all row items across all systems
    all_rows = []
    for sys_data in systems:
        all_rows.extend(sys_data.get("rows", []))

    write_rows(ws, all_rows, result, occupancy_cats=occupancy_cats)


# ── Tab dispatch ───────────────────────────────────────────────────────────────

def populate_tab(ws, tab_data: dict, result: dict, occupancy_cats: list = None):
    """Route a single tab's data to the appropriate writer."""
    if not tab_data:
        return
    if "systems" in tab_data:
        write_systems(ws, tab_data["systems"], result, occupancy_cats=occupancy_cats)
    else:
        if "static_fields" in tab_data:
            write_static_fields(ws, tab_data["static_fields"], result)
        if "rows" in tab_data:
            write_rows(ws, tab_data["rows"], result, occupancy_cats=occupancy_cats)


# ── Main function ──────────────────────────────────────────────────────────────

def populate_calculator(
    template_path: str,
    credit_code: str,
    data_json: dict,
    output_path: str,
) -> dict:
    """
    Universal LEED calculator population function.

    Args:
        template_path : Path to the USGBC template .xlsx/.xlsm file
        credit_code   : Credit code or schema ID (flexible — matched against schema)
        data_json     : Structured data extracted from project documents by Claude
        output_path   : Destination for the completed calculator file

    Returns dict: cells_written, tabs_populated, warnings, output_path, calc_name
    """
    result: dict = {
        "cells_written": 0,
        "tabs_populated": [],
        "warnings": [],
        "output_path": output_path,
        "calc_name": "",
    }

    # Step 1: Load schema
    schema = load_schema(credit_code)
    result["calc_name"] = schema.get("name", credit_code)
    print(f"  [py-calc] Schema: {schema['name']}", file=sys.stderr)

    # Step 2: Copy template → output (template is never modified)
    shutil.copy2(template_path, output_path)

    # Step 3: Open WITHOUT VBA and explicitly clear any lingering VBA binary streams.
    # Formulas and data validation survive; VBA UI buttons do not (not needed for submission).
    wb = load_workbook(output_path, keep_vba=False)
    # openpyxl sometimes carries vba_archive even with keep_vba=False — nuke it explicitly
    # so Excel doesn't complain about macro content in a .xlsx file.
    if hasattr(wb, "vba_archive"):
        wb.vba_archive = None

    # Load valid occupancy categories from Instructions sheet (for VLOOKUP snapping)
    occupancy_cats = load_occupancy_categories(wb)

    # Write project-level info across all sheets that have matching label cells
    project_info = data_json.get("project", {})
    if project_info:
        for sheet_name in wb.sheetnames:
            if sheet_name.lower() in ("instructions", "lookups", "reference", "summary"):
                continue  # skip non-data tabs
            write_static_fields(wb[sheet_name], project_info, result)

    # Steps 4–6: Process each tab listed in data_json
    tabs_data = data_json.get("tabs", {})
    for tab_name, tab_data in tabs_data.items():
        if tab_name not in wb.sheetnames:
            result["warnings"].append(f"Tab not found in workbook: {tab_name!r}")
            continue
        ws = wb[tab_name]
        populate_tab(ws, tab_data, result, occupancy_cats=occupancy_cats)
        result["tabs_populated"].append(tab_name)
        print(f"  [py-calc] Tab populated: {tab_name!r}", file=sys.stderr)

    # Step 7: Convert ArrayFormula objects to regular formula strings before saving.
    # openpyxl cannot reliably round-trip Excel array formulas in .xlsm files —
    # saving them causes Excel to show a recovery/repair dialog on next open.
    # Converting to a plain formula string loses the "array" annotation but keeps
    # the formula text, so Excel recalculates correctly without any repair needed.
    try:
        from openpyxl.worksheet.formula import ArrayFormula as _AF
        for _ws in wb.worksheets:
            for _row in _ws.iter_rows():
                for _cell in _row:
                    if isinstance(_cell.value, _AF):
                        _cell.value = _cell.value.text  # keep formula, drop array wrapper
    except (ImportError, Exception):
        pass  # older openpyxl or unexpected error — save as-is

    wb.save(output_path)
    print(
        f"  [py-calc] Saved: {Path(output_path).name} "
        f"({result['cells_written']} cells, {len(result['tabs_populated'])} tabs)",
        file=sys.stderr,
    )

    if result["warnings"]:
        for w in result["warnings"]:
            print(f"  [py-calc] warn: {w}", file=sys.stderr)

    return result


# ── CLI entry point ────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) != 5:
        print(
            f"Usage: {sys.argv[0]} <template_path> <credit_code> <data_json_path> <output_path>",
            file=sys.stderr,
        )
        sys.exit(1)

    _, tmpl, cc, djson_path, out = sys.argv

    with open(djson_path) as f:
        djson = json.load(f)

    try:
        output = populate_calculator(tmpl, cc, djson, out)
        print(json.dumps(output))
    except Exception as exc:
        print(json.dumps({"error": str(exc)}))
        import traceback
        traceback.print_exc(file=sys.stderr)
        sys.exit(1)
