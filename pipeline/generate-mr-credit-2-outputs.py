"""
Generate MR Credit 2 EPD deliverables:
  1. Populated BPDO Calculator XLSX (Materials + EPD tabs)
  2. Editable HTML version of the EPD report

Run: python3 pipeline/generate-mr-credit-2-outputs.py
"""

import shutil
import os
import re
import warnings
warnings.filterwarnings("ignore")

import openpyxl
from openpyxl.styles import PatternFill, Font

# ── Paths ─────────────────────────────────────────────────────────────────────
CALC_SRC = (
    "/Users/kelsey/Desktop/program automation "
    "/latest program versions/leed/leed calculators v4.1 nc"
    "/v4 1 Bldg products calculator_v05_07242023.xlsm"
)
OUT_CALC  = os.path.join(os.path.dirname(__file__), "output", "mr-credit-2-bpdo-calculator.xlsx")
OUT_HTML  = os.path.join(os.path.dirname(__file__), "output", "mr-credit-2-epd.html")
OUT_EDIT  = os.path.join(os.path.dirname(__file__), "output", "mr-credit-2-epd-editable.html")

# ── Product data ───────────────────────────────────────────────────────────────
# Fields: (mat_type, csi, product_name, description, manufacturer, epd_type, program_op, gbci_id)
# epd_type must match exact dropdown values in the calculator
PRODUCTS = [
    ("Ready-Mix Concrete",    "03 30 00", "Ready-Mix Concrete 4000 PSI",           "4000 PSI mix",              "CEMEX",              "Industry-wide/generic EPD",           "NRMCA / NSF",             "EPD10080"),
    ("Reinforcing Steel",     "03 20 00", "Reinforcing Steel Rebar Grade 60",       "Grade 60",                  "Nucor Steel",        "Product Specific Type III External EPD","UL Environment",         "EPD-NUCOR-UL-4790372675.101.1"),
    ("Masonry",               "04 22 00", 'Concrete Masonry Units 8"',              'Standard 8" CMU',           "Oldcastle APG",      "Product Specific Type III External EPD","ASTM International",     "ASTM Oldcastle APG CMU EPD"),
    ("Structural Steel",      "05 12 00", "Structural Steel Wide Flange W-Series",  "W-series sections",         "Nucor Steel",        "Product Specific Type III External EPD","SCS Global Services",    "SCS-EPD-10312"),
    ("Steel Deck",            "05 31 00", "Steel Roof Deck 1.5B 22ga",              "1.5B roof deck",            "Vulcraft",           "Product Specific Type III External EPD","SCS Global Services",    "SCS-EPD-09144"),
    ("Sheathing",             "06 16 00", "Exterior Gypsum Sheathing",              "DensGlass Gold",            "Georgia-Pacific",    "Product Specific Type III External EPD","NSF International",      'DensGlass 5/8" EPD'),
    ("Insulation",            "07 21 00", "Batt Insulation R-19",                   "EcoTouch PINK FIBERGLAS",   "Owens Corning",      "Product Specific Type III External EPD","UL Environment",         "Pub. No. 10023059"),
    ("Insulation",            "07 22 00", "Rigid Insulation XPS",                   "STYROFOAM Square Edge",     "Dow / DuPont",       "Product Specific Type III External EPD","ASTM International",     "STYROFOAM XPS EPD"),
    ("Roofing",               "07 54 00", "Single-Ply Roofing Membrane",            "Sarnafil G 410",            "Sika Sarnafil",      "Product Specific Type III External EPD","ASTM International",     "Sarnafil G 410 C-to-G EPD"),
    ("Sealants",              "07 92 00", "Joint Sealant Polyurethane",             "Sikaflex-1a",               "Sika",               "Product Specific Type III External EPD","ASTM International",     "Doc. 454"),
    ("Curtainwall",           "08 44 00", "Aluminum Curtainwall",                   "1600 Wall System",          "Kawneer",            "Product Specific Type III External EPD","UL Environment",         "EPD4789733794.108.1"),
    ("Doors & Frames",        "08 11 00", 'Hollow Metal Doors & Frames 18ga',       "18ga HM door & frame",      "Ceco Door",          "Industry-wide/generic EPD",           "SCS Global Services",    "SCS-EPD-10116"),
    ("Gypsum Board",          "09 29 00", 'Gypsum Wallboard 5/8"',                  "Sheetrock Brand",           "USG",                "Product Specific Type III External EPD","NSF International",      "Doc 676 / Doc 906"),
    ("Ceilings",              "09 51 00", "Acoustic Ceiling Tile 24x24",            "Ultima High-NRC",           "Armstrong",          "Product Specific Type III External EPD","ICC-ES / UL",            "ICC-ES EP104"),
    ("Flooring",              "09 68 00", "Carpet Tile",                            "Net Effect collection",     "Interface",          "Product Specific Type III External EPD","NSF International",      "Interface EPD Program"),
    ("Flooring",              "09 65 00", "Resilient Flooring LVT",                 "iD Inspiration",            "Tarkett",            "Product Specific Type III External EPD","International EPD System","iD Inspiration EPD"),
    ("Paints & Coatings",     "09 91 00", "Interior Paint",                         "Harmony Interior Latex",    "Sherwin-Williams",   "Product Specific Type III External EPD","NSF International",      "NSF EPD10546"),
    ("Paints & Coatings",     "09 91 00", "Exterior Masonry Coating",               "Loxon XP",                  "Sherwin-Williams",   "Product Specific Type III External EPD","NSF International",      "NSF EPD10800"),
    ("Paints & Coatings",     "09 96 00", "Epoxy Floor Coating",                    "ArmorSeal 1000 HS",         "Sherwin-Williams",   "Product Specific Type III External EPD","NSF International",      "NSF EPD10272"),
    ("Toilet Partitions",     "10 21 00", "Toilet Partitions Phenolic",             "B-904 Series",              "Bobrick",            "",                                    "",                       ""),
    ("Waterproofing",         "07 13 00", "Below-Grade Waterproofing",              "MasterSeal 588",            "BASF",               "",                                    "",                       ""),
    ("Tile",                  "09 30 00", "Porcelain Tile Unglazed",                "Benchmark unglazed",        "Dal-Tile",           "Industry-wide/generic EPD",           "UL Environment (TCNA)",  "TCNA 2020 IW EPD"),
    ("High Pressure Laminate","06 40 00", "High Pressure Laminate",                 "Standard HPL",              "Wilsonart",          "Product Specific Type III External EPD","SCS Global Services",    "SCS-EPD-08043"),
    ("Metal Roofing",         "07 41 00", "Standing Seam Metal Roof Panels",        "PAC-CLAD Tite-Loc Plus",    "Petersen Aluminum",  "Product Specific Type III External EPD","UL Environment",         "Petersen EPD 2020"),
    ("Exterior Cladding",     "04 43 00", "Natural Stone Exterior Cladding",        "Georgia Marble",            "Polycor",            "Product Specific Type III External EPD","Sustainable Minds",      "EPD_Company-Wide_Marble-Facades"),
]

# ── 1. Populate BPDO Calculator ───────────────────────────────────────────────
def populate_calculator():
    # Load without keep_vba — sheet formulas (cross-tab references, weight calc) are
    # stored in worksheet XML and are preserved. Only VBA macros are dropped, which
    # is fine for a data-entry calculator. This produces a clean, valid .xlsx.
    wb = openpyxl.load_workbook(CALC_SRC, keep_vba=False)
    mat = wb["Materials"]
    epd = wb["Environ. Product Declarations"]

    # Yellow fill to highlight auto-populated cells
    yellow = PatternFill(start_color="FFFBCC", end_color="FFFBCC", fill_type="solid")
    green  = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")

    # Unmerge any data-row merged cells that would block writes
    for sheet in [mat, epd]:
        to_unmerge = [str(r) for r in sheet.merged_cells.ranges if r.min_row >= 9]
        for rng in to_unmerge:
            sheet.unmerge_cells(rng)

    for i, (mat_type, csi, prod_name, desc, mfr, epd_type, prog_op, gbci_id) in enumerate(PRODUCTS):
        row = 9 + i

        # Materials tab: B=mat_type, C=csi, D=prod_name, E=desc, F=mfr, G=cost (blank)
        mat.cell(row=row, column=2).value = mat_type
        mat.cell(row=row, column=3).value = csi
        mat.cell(row=row, column=4).value = prod_name
        mat.cell(row=row, column=5).value = desc
        mat.cell(row=row, column=6).value = mfr
        # col G (cost) left blank — owner provides from Schedule of Values

        for col in [2, 3, 4, 5, 6]:
            mat.cell(row=row, column=col).fill = yellow

        # EPD tab: F=gbci_id, I=prog_op, J=epd_type
        if gbci_id:
            epd.cell(row=row, column=6).value  = gbci_id
            epd.cell(row=row, column=6).fill   = green
        if prog_op:
            epd.cell(row=row, column=9).value  = prog_op
            epd.cell(row=row, column=9).fill   = green
        if epd_type:
            epd.cell(row=row, column=10).value = epd_type
            epd.cell(row=row, column=10).fill  = green

    os.makedirs(os.path.dirname(OUT_CALC), exist_ok=True)
    # Save as xlsx (strips VBA, but openpyxl can't round-trip xlsm reliably)
    wb.save(OUT_CALC)
    print(f"  ✓ Saved calculator: {OUT_CALC}")


# ── 2. Generate editable HTML ─────────────────────────────────────────────────
EDIT_BANNER = """<div class="certifyai-edit-banner" style="background:#abcde8;color:#2b4044;padding:12px 16px;font-size:13px;font-family:Arial,Helvetica,sans-serif;display:flex;align-items:center;gap:16px;position:sticky;top:0;z-index:9999;box-shadow:0 2px 4px rgba(0,0,0,0.12);">
  <span style="flex:1;">This document is editable. Click any text to edit it directly. When finished, use <strong>File &rarr; Print &rarr; Save as PDF</strong> in your browser to save your edited version. Your edits are saved on your computer only &mdash; nothing is sent to the server.</span>
  <button onclick="window.print()" style="background:#327cb9;color:white;border:none;padding:8px 16px;border-radius:4px;font-size:13px;cursor:pointer;white-space:nowrap;font-family:inherit;font-weight:600;">Print to PDF</button>
</div>
<style>
  @media print { .certifyai-edit-banner { display: none !important; } }
  [contenteditable="true"]:focus { outline: 2px solid #327cb9; outline-offset: 1px; border-radius: 2px; background: #f0f6fc; }
  [contenteditable="true"] { cursor: text; min-height: 1em; }
</style>"""


def make_editable(html: str) -> str:
    # Insert banner right after <body>
    html = re.sub(r"(<body[^>]*>)", r"\1\n" + EDIT_BANNER, html, count=1)

    # Make .field-value spans editable
    html = re.sub(
        r'(<span class="field-value">)(.*?)(</span>)',
        r'\1<span contenteditable="true">\2</span>\3',
        html,
        flags=re.DOTALL,
    )

    # Make table data cells editable (td only, not th)
    # Skip formula/badge cells — make plain text cells editable
    html = re.sub(
        r'(<td(?:\s[^>]*)?>)(<a |<span class="(?:tag-|note-|badge-|num))',
        r'\1\2',
        html,
    )
    # Simple text-only td values get contenteditable
    html = re.sub(
        r'(<td(?:\s+class="num"[^>]*)?>)(\d[\d.]*)(</td>)',
        r'\1<span contenteditable="true">\2</span>\3',
        html,
    )

    return html


def generate_editable():
    with open(OUT_HTML, "r", encoding="utf-8") as f:
        html = f.read()
    editable = make_editable(html)
    with open(OUT_EDIT, "w", encoding="utf-8") as f:
        f.write(editable)
    print(f"  ✓ Saved editable HTML: {OUT_EDIT}")


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("Generating MR Credit 2 outputs...")
    populate_calculator()
    generate_editable()
    print("Done.")
